import openpyxl

import unittest
from utlis.handle_path import *
from utlis.handle_excel import Handle_Excel
from library.ddt import ddt,data  #从ddt模块中导入ddt和data装饰器
from utlis.handle_conf import conf
from utlis.handle_requests import Send_Requests
import os

class Readexcel():
    def __init__(self,filname,sheetname):
        self.filname=filname
        self.sheetname=sheetname       #创建初始化函数

    def open(self):
        #打开数据表
        self.workbook=openpyxl.load_workbook(self.filname)  #打开工作簿
        self.sheetname=self.workbook[self.sheetname]#定位具体的sheet表单

    def save(self):
        #保存数据
        self.workbook.save(self.filname)

    def write(self,row,column,value):
        #excel写入数据
        self.sheetname.cell(row=row,column=column,value=value)
        self.save()

    def read(self):
        #读取数据，遍历添加到字典里
        self.open()
        max_row=self.sheetname.max_row
        max_column=self.sheetname.max_column
        datas=[]
        for row in range(1,max_row+1):
            data_list=[]
            for column in range(1,max_column+1):
                data=self.sheetname.cell(row=row,column=column).value
                data_list.append(data)
            datas.append(data_list)
        cases=[]
        tile=datas[0]
        for element in datas[1:]:
            new_data=dict(zip(tile,element))
            cases.append(new_data)
        return cases

# if __name__ == '__main__':
    # path = os.path.join(data_path, 'apicases.xlsx')
    # a=Readexcel(path,"login")
    # print(a.read())





