import openpyxl#打开Excel的工具
import os
from utlis.handle_path import *

class Handle_Excel(object):
    '''封装一个读取Excel表格的工具类'''

    def __init__(self,filename,sheet_name):   #构建初始化函数
        self.filename = filename   #excel文件名#把传进来的形式参数赋予到实列变量self.filename
        self.sheet_name = sheet_name#Excel的表单名

    def open(self):
        '''封装一个打开Excel表格的工具类'''

        # 通过openpyxl模块调用load_workbook()函数读取filename文件中的内容
        self.wb = openpyxl.load_workbook(self.filename)
#通过self.wb这个Excel文件对象读取对应的sheet_name页面
        self.sh = self.wb[self.sheet_name]

    def read_data(self):
        #封装一个读取excel表格的工具
        self.open()
        datas = list(self.sh.rows)#把每一行的元祖数据放入list列表中
        # print(datas)
        #用列表解析式
        title = [i.value for i in datas[0]]  #datas[0]为第一行数据的对象
        #创建一个空用例的列表，用来接收所有的测试
        cases = []
        for j in datas[1:]: #对2行以后进行for循环
            values = [k.value for k in j]#然后通过对j进行遍历，用k.value获取每个用例对象当中的value值放入到一个列表中
            case = dict(zip(title,values))#把title列表和values每个用例一一对应打包放入在一个字典中
            cases.append(case)#吧每个用例追加到列表中，把所有的用例进行返回
        return cases
    def write_excel(self,row,column,value=None):  #row=行  columu=列
        '''封装一个往Excel表格里面写入测试结果的工具方法'''
        # 打开Excel表格
        self.open()
        #往固定的row行和column列写入values数据
        self.sh.cell(row,column,value)
        #保存Excel表格并关闭
        self.wb.save(self.filename)
path = os.path.join(data_path,'apicases.xlsx')
read_excel = Handle_Excel(path,'login')
if __name__ == '__main__':
    path = os.path.join(data_path, 'apicases.xlsx')
    read_excel = Handle_Excel(path, 'login')
    print(read_excel.read_data())
   # print(read_excel.read_data())
   #  read_excel.write_excel(2,8,'不通过')







# # coding=utf-8
#
# import openpyxl
# import os
# from utlis.handle_path import *
#
#
# class Handle_Excel(object):
#     '''封装一个读取Excel表格的工具类'''
#
#     def __init__(self,filename,sheet_name):
#         self.filename = filename
#         self.sheet_name = sheet_name
#
#     def open(self):
#         '''封装一个打开Excel表格的工具类'''
#
#         #通过openpyxl模块调用load_workbook()函数读取filename文件中的内容
#         self.wb = openpyxl.load_workbook(self.filename)
#
#         #通过self.wb这个Excel文件对象读取对应的sheet_name页面
#         self.sh = self.wb[self.sheet_name]
#
#     def read_data(self):
#         '''封装一个读取Excel表格的工具'''
#         self.open()
#         datas = list(self.sh.rows) #把每一行的元组数据放入list列表中
#         # print(datas) data是取所有行数据的对象   datas[0]为第一行数据的对象
#         #用列表解析式
#         title = [i.value for i in datas[0]]
#         #print(title)   #['case_id', 'interface', 'title', 'method', 'url', 'data', 'excepted', 'result']
#         #创建一个空的列表,用来接收所有的测试用例
#         cases = []
#         for j in datas[1:]:  #对2/3/4/5行进行for循环
#             values = [k.value for k in j]  #然后通过对j进行遍历,用k.value获取每个用例对象当中的value值放入到一个列表中
#             case = dict(zip(title,values)) #把title 列表和values每个用例一一对应打包放在一个字典当中
#             cases.append(case)  # 把每个用例追加到列表中
#         return cases            #把所有的用例进行返回
#
#     def write_excel(self,row,column,value=None):
#         '''封装一个往Excel表格里面写入测试结果的工具方法'''
#         #打开Excel表格
#         self.open()
#         #往固定的row行和column列写入value数据
#         self.sh.cell(row,column,value)
#         #保存Excel表格并且关闭
#         self.wb.save(self.filename)
#
#
# path = os.path.join(data_path,'apicases.xlsx')
# read_excel = Handle_Excel(path,'login')
# # if __name__ == '__main__':
# #
# #  print(read_excel.read_data())
# #  read_excel.write_excel(2,8,'通过')
